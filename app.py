import streamlit as st
from streamlit_gsheets import GSheetsConnection
from streamlit_calendar import calendar
import pandas as pd
from datetime import datetime
import io
import os

# PDF/Excel Libraries
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ================= CONFIGURATION =================
st.set_page_config(page_title="Interview Scheduler", layout="wide", page_icon="🕒")
conn = st.connection("gsheets", type=GSheetsConnection)

# ================= CONSTANTS =================
# Time range: 09:00 AM to 12:00 AM (24:00)
TIME_SLOTS = []
for h in range(9, 24):  # 9 to 23
    for m in (0, 30):
        TIME_SLOTS.append(f"{h:02d}:{m:02d}")

# 科目清單
SUBJECT_OPTIONS = ["中文", "英文", "數學", "生物", "地理", "中史", "歷史", "物理", "化學"]

# ================= DATA FUNCTIONS =================
def clean_dataframe(df):
    """清理資料並標準化格式"""
    # 轉為字串以處理 NaN
    df = df.astype(str)
    
    # 清理無效字串
    for col in df.columns:
        if col not in ['LastUpdated']:
            df[col] = df[col].replace(['NaT', 'nan', 'None', '<NA>'], '')
            
    # 處理日期格式
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.strftime('%Y-%m-%d')
    
    # 處理時間格式 (相容 HH:MM:SS 與 HH:MM)
    df['Time'] = pd.to_datetime(df['Time'], format='%H:%M:00', errors='coerce').fillna(
                 pd.to_datetime(df['Time'], format='%H:%M', errors='coerce')
                 ).dt.strftime('%H:%M')
                 
    df = df.fillna("")
    
    # 處理 LastUpdated
    if 'LastUpdated' not in df.columns:
        df['LastUpdated'] = pd.NaT
    df['LastUpdated'] = pd.to_datetime(df['LastUpdated'], errors='coerce')
    
    return df

def load_data_from_google():
    try:
        df = conn.read(worksheet="Sheet1", ttl=0)
        if df.empty:
            return pd.DataFrame(columns=["Name", "ID", "Date", "Time", "Notes", "LastUpdated"])
        return clean_dataframe(df)
    except Exception as e:
        if "429" in str(e):
            st.error("⚠️ Too many requests! Please wait 1 minute before refreshing.")
        else:
            st.error(f"Database Error: {e}")
        return pd.DataFrame(columns=["Name", "ID", "Date", "Time", "Notes", "LastUpdated"])

def initialize_session():
    if 'data' not in st.session_state:
        with st.spinner("🔄 Connecting to Cloud Database..."):
            st.session_state.data = load_data_from_google()
    
    if 'form_id' not in st.session_state:
        st.session_state.form_id = 0
    if 'last_cloud_timestamp' not in st.session_state:
        if not st.session_state.data.empty and 'LastUpdated' in st.session_state.data.columns:
             st.session_state.last_cloud_timestamp = st.session_state.data['LastUpdated'].max()
        else:
             st.session_state.last_cloud_timestamp = None

def refresh_data(force=False):
    st.cache_data.clear()
    new_data = load_data_from_google()
    st.session_state.data = new_data
    if not new_data.empty and 'LastUpdated' in new_data.columns:
        max_ts = new_data['LastUpdated'].max()
        st.session_state.last_cloud_timestamp = max_ts
    st.toast("Data refreshed from Cloud", icon="🔄")
    if force:
        st.rerun()

# ================= CONFLICT DETECTION & SAVE =================
def save_with_conflict_detection(new_df):
    try:
        latest_cloud = load_data_from_google()
        
        cloud_latest_ts = pd.NaT
        if not latest_cloud.empty and 'LastUpdated' in latest_cloud.columns:
            cloud_latest_ts = latest_cloud['LastUpdated'].max()
            if pd.notna(cloud_latest_ts):
                cloud_latest_ts = cloud_latest_ts.tz_localize(None)

        user_latest_ts = st.session_state.last_cloud_timestamp
        if pd.notna(user_latest_ts):
            user_latest_ts = user_latest_ts.tz_localize(None)

        if pd.notna(user_latest_ts) and pd.notna(cloud_latest_ts) and cloud_latest_ts > user_latest_ts:
            st.error("⚠️ 儲存失敗：檢測到雲端資料已被其他人修改！")
            st.write(f"雲端最新: {cloud_latest_ts}")
            st.write(f"本地基準: {user_latest_ts}")
            
            col1, col2 = st.columns(2)
            if col1.button("🔄 放棄修改並重新載入"):
                refresh_data(force=True)
                return
            if col2.button("⚠️ 強制覆蓋 (可能遺失他人修改)", type="primary"):
                pass
            else:
                st.stop()
        
        current_time = pd.Timestamp.now()
        
        # 重置索引
        clean_df = clean_dataframe(new_df.copy()).reset_index(drop=True)
        
        clean_df['LastUpdated'] = current_time
        
        # 轉換為字串以上傳
        upload_df = clean_df.copy()
        upload_df['LastUpdated'] = upload_df['LastUpdated'].dt.strftime('%Y-%m-%d %H:%M:%S')

        conn.update(worksheet="Sheet1", data=upload_df)
        
        st.session_state.data = clean_df
        st.session_state.last_cloud_timestamp = current_time
        
        st.success("✅ 儲存成功！雲端資料已更新。")
        st.rerun()
        
    except Exception as e:
        if "429" in str(e):
            st.error("⚠️ 儲存失敗：請求過多，請等待 60 秒後再試。")
        else:
            st.error(f"儲存失敗：{e}")
            st.write("詳細錯誤資訊：", str(e))

# ================= EXPORT FUNCTIONS =================
def generate_visual_pdf(df):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    elements = []
    font_name = "Helvetica"
    try:
        if os.path.exists("font.ttf"):
            pdfmetrics.registerFont(TTFont('CustomChinese', 'font.ttf'))
            font_name = 'CustomChinese'
        elif os.path.exists("font.otf"):
            pdfmetrics.registerFont(TTFont('CustomChinese', 'font.otf'))
            font_name = 'CustomChinese'
    except Exception as e:
        print(f"Font loading error: {e}")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ChineseTitle', parent=styles['Heading1'], fontName=font_name, fontSize=16, leading=20)
    cell_style = ParagraphStyle('ChineseCell', parent=styles['Normal'], fontName=font_name, fontSize=9, leading=11)
    
    df['dt'] = pd.to_datetime(df['Date'] + " " + df['Time'], errors='coerce')
    df = df.dropna(subset=['dt'])
    months = sorted(df['dt'].dt.to_period('M').unique())
    import calendar as py_calendar
    cal = py_calendar.Calendar(firstweekday=6)

    for period in months:
        year, month = period.year, period.month
        elements.append(Paragraph(f"<b>{period.strftime('%B %Y')}</b>", title_style))
        elements.append(Spacer(1, 10))
        
        month_cal = cal.monthdayscalendar(year, month)
        table_data = [["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]]
        row_heights = [20]

        for week in month_cal:
            row_cells = []
            max_entries = 0
            for day in week:
                if day == 0:
                    row_cells.append("")
                else:
                    day_str = f"{year}-{month:02d}-{day:02d}"
                    day_data = df[df['Date'] == day_str].sort_values('Time')
                    cell_text = f"<b>{day}</b>"
                    if not day_data.empty:
                        lines = [f"{r['Name']}\n{r['Time']}" for _, r in day_data.iterrows()]
                        cell_text += "\n\n" + "\n".join(lines)
                        max_entries = max(max_entries, len(day_data))
                    row_cells.append(Paragraph(cell_text.replace("\n", "<br/>"), cell_style))
            table_data.append(row_cells)
            row_heights.append(40 + (max_entries * 25))

        table = Table(table_data, colWidths=[110]*7, rowHeights=row_heights)
        table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('FONTNAME', (0,0), (-1,-1), font_name),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_visual_excel(df):
    wb = Workbook()
    wb.remove(wb.active)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    import calendar as py_calendar
    cal = py_calendar.Calendar(firstweekday=6)
    
    df['dt'] = pd.to_datetime(df['Date'] + " " + df['Time'], errors='coerce')
    months = sorted(df['dt'].dt.to_period('M').dropna().unique())

    for period in months:
        ws = wb.create_sheet(f"{period.year}-{period.month:02d}")
        ws.merge_cells("A1:G1")
        ws["A1"] = f"{period.strftime('%B %Y')}"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        
        for i, d in enumerate(["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], 1):
            c = ws.cell(2, i, d)
            c.fill = PatternFill("solid", fgColor="DDDDDD")
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64+i)].width = 20

        row_num = 3
        for week in cal.monthdayscalendar(period.year, period.month):
            max_h = 1
            for col_idx, day in enumerate(week, 1):
                c = ws.cell(row_num, col_idx)
                c.border = thin
                c.alignment = align
                if day != 0:
                    day_str = f"{period.year}-{period.month:02d}-{day:02d}"
                    day_data = df[df['Date'] == day_str].sort_values('Time')
                    val = f"{day}\n"
                    if not day_data.empty:
                        lines = [f"{r['Name']} ({r['Time']})" for _, r in day_data.iterrows()]
                        val += "\n".join(lines)
                        max_h = max(max_h, len(lines)+1)
                    c.value = val
            ws.row_dimensions[row_num].height = max(50, max_h * 15)
            row_num += 1
            
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ================= MAIN APP LOGIC =================
initialize_session()
df = st.session_state.data

st.title("🕒 Cloud Scheduler (多人協作優化版)")

if not df.empty and 'LastUpdated' in df.columns:
    max_ts = df['LastUpdated'].max()
    if pd.notna(max_ts):
        st.caption(f"📅 Last Updated: {max_ts.strftime('%Y-%m-%d %H:%M:%S')}")

st.warning("⚠️ **For multi-user editing, please click '🔄 Force Sync' first to avoid overwriting others' data!**")

if st.button("🔄 Force Sync from Cloud"):
    refresh_data(force=True)

tab1, tab2, tab3 = st.tabs(["📅 Calendar View", "📝 List & Edit", "⚙️ Export & Import"])

# --- TAB 1: CALENDAR ---
with tab1:
    if not df.empty:
        df_cal = df.reset_index(drop=True)
        events = []
        for index, row in df_cal.iterrows():
            if row['Date'] and row['Time'] and len(str(row['Date'])) > 0 and len(str(row['Time'])) > 0:
                try:
                    start_iso = f"{row['Date']}T{row['Time']}"
                    events.append({
                        "id": str(index),
                        "title": row['Name'],
                        "start": start_iso,
                        "extendedProps": {"description": f"ID: {row['ID']} | Notes: {row['Notes']}"}
                    })
                except:
                    continue
        
        calendar(events=events, options={
            "initialView": "dayGridMonth",
            "height": "750px",
            "headerToolbar": {"left": "prev,next today", "center": "title", "right": "dayGridMonth,listMonth"},
            "eventTimeFormat": {"hour": "2-digit", "minute": "2-digit", "hour12": False}
        }, key="main_calendar")
    else:
        st.info("No data available.")

# --- TAB 2: EDIT ---
with tab2:
    c1, c2 = st.columns([1, 2])
    with c1:
        st.subheader("Add Booking")
        
        limit = st.number_input("Max People per Slot (0 = Unlimited)", min_value=0, value=0)

        with st.form("add", clear_on_submit=False):
            form_id = st.session_state.form_id
            
            # 使用下拉選單選擇科目
            name = st.selectbox("科目 (Subject)", SUBJECT_OPTIONS, key=f"name_{form_id}")
            
            c_id = st.text_input("ID / Class", key=f"id_{form_id}")
            d = st.date_input("Date", min_value=datetime.today(), key=f"date_{form_id}")
            t_str = st.selectbox("Time", TIME_SLOTS, key=f"time_{form_id}")
            notes = st.text_area("Notes", key=f"notes_{form_id}")
            
            if st.form_submit_button("Save"):
                if not name:
                    st.error("Name/Subject required")
                else:
                    limit_reached = False
                    if limit > 0:
                        check_date = d.strftime("%Y-%m-%d")
                        existing_count = len(df[(df['Date'] == check_date) & (df['Time'] == t_str)])
                        if existing_count >= limit:
                            limit_reached = True
                            st.error(f"⛔ Slot {check_date} {t_str} is FULL! ({existing_count}/{limit})")
                    
                    if not limit_reached:
                        new_row = pd.DataFrame([{
                            "Name": name, "ID": c_id, "Date": d.strftime("%Y-%m-%d"),
                            "Time": t_str, "Notes": notes, "LastUpdated": pd.NaT
                        }])
                        new_df = pd.concat([df, new_row], ignore_index=True)
                        save_with_conflict_detection(new_df)
                        st.session_state.form_id += 1

    with c2:
        st.subheader("Edit Grid")
        st.caption("Double-click to edit. Select rows & Delete key to remove.")
        
        edit_in = df.copy()
        edit_in["Date"] = pd.to_datetime(edit_in["Date"], errors='coerce').dt.date
        edit_in["Time"] = pd.to_datetime(edit_in["Time"], format='%H:%M', errors='coerce').dt.time
        
        out = st.data_editor(
            edit_in,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                # 設定 Name 欄位在表格編輯時也顯示為下拉選單
                "Name": st.column_config.SelectboxColumn(
                    "科目 (Subject)",
                    help="選擇科目",
                    width="medium",
                    options=SUBJECT_OPTIONS,
                    required=True,
                ),
                "Time": st.column_config.TimeColumn("Time", format="HH:mm", step=1800),
                "LastUpdated": None  # Hide
            }
        )
        
        if st.button("💾 Save Changes to Cloud", type="primary"):
            clean_out = out.copy()
            # 修正：確保在呼叫 strftime 前檢查是否為 NaT
            clean_out['Date'] = clean_out['Date'].apply(
                lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and hasattr(x, 'strftime') else (str(x) if pd.notna(x) and str(x) != 'NaT' else '')
            )
            clean_out['Time'] = clean_out['Time'].apply(
                lambda x: x.strftime('%H:%M') if pd.notna(x) and hasattr(x, 'strftime') else (str(x) if pd.notna(x) and str(x) != 'NaT' else '')
            )
            save_with_conflict_detection(clean_out)

# --- TAB 3: EXPORT ---
with tab3:
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### Visual Reports")
        st.write("If Chinese characters are missing in PDF, make sure `font.ttf` is uploaded.")
        if not df.empty:
            st.download_button("📄 PDF Calendar", generate_visual_pdf(df), "cal.pdf", "application/pdf")
            st.download_button("📊 Excel Calendar", generate_visual_excel(df), "cal.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    with col2:
        st.markdown("### Import Backup")
        up = st.file_uploader("CSV", type="csv")
        if up and st.button("Import"):
            try:
                imp = pd.read_csv(up).fillna("")
                if 'Name' in imp.columns:
                    # Add LastUpdated on import
                    imp['LastUpdated'] = pd.NaT
                    save_with_conflict_detection(pd.concat([df, imp], ignore_index=True))
                    st.success("Imported!")
                    st.rerun()
            except:
                st.error("Invalid CSV")
